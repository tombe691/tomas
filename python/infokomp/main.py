from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import time

# Ladda ner chromedriver och ange sökvägen här
PATH = "C:\Program Files (x86)\chromedriver.exe"
option = webdriver.ChromeOptions()
# Om du inte använder brave kan du ta bort den här raden och options från raden under
# Brave är dock snabbare än chrome så om det blir något fel som säger att elementet inte
# kan hittas så beror det nog på att browsern inte hinner med.
option.binary_location = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
driver = webdriver.Chrome(PATH, options=option)

driver.get("https://infokomp.itslearning.com/")

now = datetime.now()
dt_string = now.strftime(r"%Y-%m-%d %H-%M-%S")

assign_name = "Uppgifter " + dt_string
week_name = "Närvaro " + dt_string

# Ändra till sökväg och namn på den excelfil du vill spara datan på.
# Kom ihåg att stänga excelfilen innan du kör
excel_path = r"C:\Tomas\infokomp\Data23.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
except:
    wb = Workbook()

sh = wb.create_sheet(assign_name)
nsh = wb.create_sheet(week_name)

sh.column_dimensions['A'].width = 25
sh.column_dimensions['B'].width = 25
sh.column_dimensions['C'].width = 17
sh.column_dimensions['D'].width = 90
sh['A1'] = "Uppgift"
sh['B1'] = "Elev"
sh['C1'] = "Datum"
sh['D1'] = "URL"

nsh.column_dimensions['A'].width = 25
nsh.column_dimensions['B'].width = 25
nsh.column_dimensions['C'].width = 17

nsh['A1'] = "Elev"
nsh['B1'] = "Datum"
nsh['C1'] = "Bedömning"

login = driver.find_element_by_id("ctl00_ContentPlaceHolder1_Username_input")
login.send_keys("TomasBe")

password = driver.find_element_by_id("ctl00_ContentPlaceHolder1_Password_input")
password.send_keys("rH$V;5U9")

button = driver.find_element_by_id("ctl00_ContentPlaceHolder1_nativeLoginButton")
button.click()

iframe = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentAreaIframe"]')
driver.switch_to.frame(iframe)

wait = WebDriverWait(driver, 10)
element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'coursecard__main')))
courses_length = len(driver.find_elements_by_class_name("coursecard__main"))
print(courses_length)
driver.implicitly_wait(10)

courselist = []
datelist = []
exlist = []

for i in range(courses_length):
    courses = driver.find_elements_by_class_name("coursecard__main")
    course = courses[i]

    if (course.text.__contains__("Följ upp uppgifter")):
        courselist.append(i)
    elif (course.text.__contains__("Vecka")):
        datelist.append(i)
    elif (course.text.__contains__("Slutexamination")):
        exlist.append(i)


driver.switch_to.parent_frame()

for i in range(len(courselist)):

    iframe = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentAreaIframe"]')
    driver.switch_to.frame(iframe)
    courses = driver.find_elements_by_class_name("coursecard__main")
    course = courses[courselist[i]]
    course.click()

    driver.fullscreen_window()
    iframe = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentAreaIframe"]')
    driver.switch_to.frame(iframe)

    element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'tasklist2-task')))
    assignments_length = len(driver.find_elements_by_class_name("tasklist2-task"))
    print(assignments_length)
    assignlist = []
    weeklist = []
    weekindex = 0

    for a in range(assignments_length):
        assignments = driver.find_elements_by_class_name("tasklist2-task")
        assign = assignments[a]

        if (assign.text.__contains__("Uppgift")):
            assignlist.append(a)
        else:
            if (len(weeklist) <= 0):
                weekindex = a
            weeklist.append(a)
    print(len(weeklist))
    print(weekindex)
    driver.switch_to.parent_frame()

    for j in range(len(assignlist)):
        iframe = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentAreaIframe"]')
        driver.switch_to.frame(iframe)
        assignments = driver.find_elements_by_class_name("tasklist2-task")
        assignment = assignments[assignlist[j]]
        assignment.click()

        header = driver.find_element_by_id("ctl00_PageHeader_TT")
        header_text = header.text
        get_url = driver.current_url

        iframe = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder_ExtensionIframe"]')
        driver.switch_to.frame(iframe)
        uppgifter = driver.find_elements_by_class_name("h-cursor-pointer")
        names = driver.find_elements_by_class_name("as-person-reference-name")
        dates = driver.find_elements_by_class_name("as-hide-if-screen-below-767")

        for k in range(len(uppgifter)):
            uppgift = uppgifter[k]
            name = names[k]
            date = dates[k + 1]
            if (uppgift.text.__contains__("INLÄMNAD")):
                if (uppgift.text.__contains__("INTE")):
                    break
                else:
                    info = (header_text, name.text, date.text, get_url)
                    sh.append(info)
                    wb.save(excel_path)
        driver.switch_to.parent_frame()
        driver.back()

    if (len(weeklist) > 0):
        print("Inside week before for")
        for f in range(len(weeklist)):
            print("Inside week")
            iframe = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentAreaIframe"]')
            driver.switch_to.frame(iframe)
            weeks = driver.find_elements_by_class_name("tasklist2-task")
            week = weeks[weekindex]
            week.click()

            names = driver.find_elements_by_class_name("name")
            dates = driver.find_elements_by_class_name("answered")
            assessments = driver.find_elements_by_class_name("assesment")
            print(len(names))

            for g in range(len(names)):
                name = names[g]
                date = dates[g]
                assessment = assessments[g]
                print(name)

                info = (name.text, date.text, assessment.text)
                nsh.append(info)
                wb.save(excel_path)

            driver.switch_to.parent_frame()
            driver.back()

    driver.back()
    for r in range(len(datelist)):
        iframe = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentAreaIframe"]')
        driver.switch_to.frame(iframe)
        courses = driver.find_elements_by_class_name("coursecard__main")
        course = courses[datelist[r]]
        print(r)
        course.click()
driver.close()